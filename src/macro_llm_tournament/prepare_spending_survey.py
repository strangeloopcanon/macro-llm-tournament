from __future__ import annotations

import argparse
import hashlib
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_MICRODATA = PROJECT_ROOT / "work" / "spending_survey_raw" / "sce-household-spending-microdata.xlsx"
DEFAULT_CHART_DATA = PROJECT_ROOT / "work" / "spending_survey_raw" / "sce-household-spending-chart-data.xlsx"
DEFAULT_CORE_PANEL = PROJECT_ROOT / "work" / "persona_beliefs" / "sce_real_microdata.csv"
DEFAULT_OUTPUT = PROJECT_ROOT / "work" / "empirical_bridge" / "spending_belief_panel.csv"
DEFAULT_COVERAGE = PROJECT_ROOT / "work" / "empirical_bridge" / "spending_belief_panel_coverage.json"

SPENDING_PANEL_VERSION = "empirical_bridge_spending_panel_v3"
EXPECTED_SPENDING_COLUMNS = tuple(f"qsp7dens_{index}" for index in range(1, 11))
REQUIRED_SPENDING_COLUMNS = ("userid", "date", "qsp1", "qsp2", *EXPECTED_SPENDING_COLUMNS, "qsp12n", "qsp13new")
REGRESSORS = (
    "actual_expected_inflation_1y",
    "actual_expected_real_income_growth",
    "sce_question_unemployment_higher_prob",
)
FIT_WAVES = (202004, 202008, 202012, 202104, 202108, 202112, 202204, 202208, 202304, 202308)
INTERNAL_CHECK_WAVES = (202212,)
VALIDATION_WAVES = (202312, 202404, 202408)
_GENERALIZED_BETA_CACHE: dict[tuple[float, ...], float] = {}


@dataclass(frozen=True)
class ChartReproduction:
    series: str
    max_abs_error_pp: float
    passed: bool
    rows: list[dict[str, Any]]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the empirical-bridge spending/belief panel.")
    parser.add_argument("--microdata-xlsx", type=Path, default=DEFAULT_MICRODATA)
    parser.add_argument("--chart-data-xlsx", type=Path, default=DEFAULT_CHART_DATA)
    parser.add_argument("--core-panel-csv", type=Path, default=DEFAULT_CORE_PANEL)
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--coverage-json", type=Path, default=DEFAULT_COVERAGE)
    parser.add_argument("--max-chart-error-pp", type=float, default=0.1)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = build_spending_belief_panel(
        microdata_xlsx=args.microdata_xlsx,
        chart_data_xlsx=args.chart_data_xlsx,
        core_panel_csv=args.core_panel_csv,
        output_csv=args.output_csv,
        coverage_json=args.coverage_json,
        max_chart_error_pp=args.max_chart_error_pp,
    )
    print(json.dumps(result, indent=2, sort_keys=True))
    return 0


def build_spending_belief_panel(
    *,
    microdata_xlsx: Path,
    chart_data_xlsx: Path,
    core_panel_csv: Path,
    output_csv: Path,
    coverage_json: Path,
    max_chart_error_pp: float = 0.1,
) -> dict[str, Any]:
    if not microdata_xlsx.exists():
        raise FileNotFoundError(f"Spending microdata not found: {microdata_xlsx}")
    if not chart_data_xlsx.exists():
        raise FileNotFoundError(f"Spending chart data not found: {chart_data_xlsx}")
    if not core_panel_csv.exists():
        raise FileNotFoundError(f"Core SCE belief panel not found: {core_panel_csv}")

    spending = normalize_spending_microdata(read_spending_workbook(microdata_xlsx))
    core = normalize_core_panel(pd.read_csv(core_panel_csv))
    chart = read_chart_data(chart_data_xlsx)
    expected_check = verify_expected_spending_chart(spending, chart, core, max_abs_error_pp=max_chart_error_pp)
    reported_check = verify_reported_spending_chart(spending, core, chart, max_abs_error_pp=max_chart_error_pp)
    panel, join_report = join_spending_to_core_panel(spending, core)
    if panel.empty:
        raise ValueError("No matched spending/belief rows after the three-month core-panel join")

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    coverage_json.parent.mkdir(parents=True, exist_ok=True)
    panel.to_csv(output_csv, index=False)
    coverage = {
        "schema_version": SPENDING_PANEL_VERSION,
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "status": "ok",
        "source_files": {
            "spending_microdata_path": _safe_relative(microdata_xlsx),
            "spending_microdata_sha256": file_sha256(microdata_xlsx),
            "chart_data_path": _safe_relative(chart_data_xlsx),
            "chart_data_sha256": file_sha256(chart_data_xlsx),
            "core_panel_path": _safe_relative(core_panel_csv),
            "core_panel_sha256": file_sha256(core_panel_csv),
        },
        "output_csv": _safe_relative(output_csv),
        "output_csv_sha256": file_sha256(output_csv),
        "question_mapping": {
            "reported_total_spending_growth_pct": "QSP1 sign x QSP2 magnitude",
            "expected_total_spending_growth_pct": "QSP7dens_1..10 density response, generalized beta mean",
            "windfall_income_gain": "QSP12n / QSP12a_1..3",
            "windfall_income_loss": "QSP13new / QSP13a_1..3",
        },
        "chart_reproduction": {
            "expected_total_spending_growth_pct": expected_check.__dict__,
            "reported_total_spending_growth_pct": reported_check.__dict__,
            "gate_passed": bool(expected_check.passed and reported_check.passed),
            "gate_tolerance_pp": float(max_chart_error_pp),
            "note": (
                "Expected-spending medians use respondent-level generalized beta means. "
                "Reported-spending medians require SCE person weights; the spending workbook has no native weight column, "
                "so same-wave core-panel weights are used and mismatches are reported fail-closed."
            ),
        },
        "join_report": join_report,
        "fit_waves": list(FIT_WAVES),
        "internal_check_waves": list(INTERNAL_CHECK_WAVES),
        "validation_waves": list(VALIDATION_WAVES),
        "row_count": int(panel.shape[0]),
        "respondent_count": int(panel["userid"].nunique()),
        "wave_count": int(panel["spending_wave"].nunique()),
        "spending_wave_min": str(panel["spending_wave"].min()),
        "spending_wave_max": str(panel["spending_wave"].max()),
    }
    if not coverage["chart_reproduction"]["gate_passed"]:
        coverage["status"] = "chart_reproduction_failed"
        write_blocker(
            "Deliverable 1 chart-data reproduction did not pass the locked +/-0.1pp gate. "
            "The question codes are present and mapped from the questionnaire/glossary, but at least one "
            "published chart median does not reproduce with public spending microdata plus core SCE weights. "
            "See work/empirical_bridge/spending_belief_panel_coverage.json for per-wave errors."
        )
    coverage_json.write_text(json.dumps(_jsonable(coverage), indent=2, sort_keys=True), encoding="utf-8")
    return {
        "output_csv": str(output_csv),
        "coverage_json": str(coverage_json),
        "row_count": int(panel.shape[0]),
        "chart_gate_passed": bool(coverage["chart_reproduction"]["gate_passed"]),
    }


def read_spending_workbook(path: Path) -> pd.DataFrame:
    try:
        import python_calamine  # noqa: F401

        frame = pd.read_excel(path, sheet_name="Data", header=1, engine="calamine")
    except ImportError:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        header_index = next(
            index for index, row in enumerate(rows[:10]) if {"userid", "date", "qsp1"}.issubset({str(value) for value in row if value is not None})
        )
        frame = pd.DataFrame(rows[header_index + 1 :], columns=[str(value) if value is not None else "" for value in rows[header_index]])
    missing = sorted(set(REQUIRED_SPENDING_COLUMNS) - set(frame.columns))
    if missing:
        raise ValueError(f"Spending workbook missing required columns: {', '.join(missing)}")
    return frame


def normalize_spending_microdata(raw: pd.DataFrame) -> pd.DataFrame:
    frame = raw.copy()
    frame["userid"] = frame["userid"].astype(str).str.replace(r"\.0$", "", regex=True)
    frame["spending_wave"] = pd.to_numeric(frame["date"], errors="coerce").astype("Int64")
    frame["spending_wave_date"] = _wave_to_date(frame["spending_wave"])
    qsp1 = pd.to_numeric(frame["qsp1"], errors="coerce")
    qsp2 = pd.to_numeric(frame["qsp2"], errors="coerce")
    frame["reported_total_spending_growth_pct"] = np.where(qsp1.eq(1), qsp2, np.where(qsp1.eq(2), -qsp2, np.nan))
    frame["expected_total_spending_growth_pct"] = frame.apply(expected_spending_growth_from_density, axis=1)
    windfall = normalize_windfall_rows(frame)
    out = pd.concat(
        [
            frame[
                [
                    "userid",
                    "spending_wave",
                    "spending_wave_date",
                    "reported_total_spending_growth_pct",
                    "expected_total_spending_growth_pct",
                ]
            ],
            windfall,
        ],
        axis=1,
    )
    out = out.dropna(subset=["userid", "spending_wave", "spending_wave_date", "expected_total_spending_growth_pct"]).copy()
    return out.reset_index(drop=True)


def normalize_windfall_rows(frame: pd.DataFrame) -> pd.DataFrame:
    gain_choice = pd.to_numeric(frame.get("qsp12n"), errors="coerce")
    loss_choice = pd.to_numeric(frame.get("qsp13new"), errors="coerce")
    out = pd.DataFrame(index=frame.index)
    for column, fallback in [("qsp12a_1", 0.0), ("qsp12a_2", 0.0), ("qsp12a_3", 0.0), ("qsp13a_1", 0.0), ("qsp13a_2", 0.0), ("qsp13a_3", 0.0)]:
        out[column] = pd.to_numeric(frame.get(column, fallback), errors="coerce")
    out.loc[gain_choice.eq(1), ["qsp12a_1", "qsp12a_2", "qsp12a_3"]] = [100.0, 0.0, 0.0]
    out.loc[gain_choice.eq(2), ["qsp12a_1", "qsp12a_2", "qsp12a_3"]] = [0.0, 100.0, 0.0]
    out.loc[gain_choice.eq(3), ["qsp12a_1", "qsp12a_2", "qsp12a_3"]] = [0.0, 0.0, 100.0]
    out.loc[loss_choice.eq(1), ["qsp13a_1", "qsp13a_2", "qsp13a_3"]] = [100.0, 0.0, 0.0]
    out.loc[loss_choice.eq(2), ["qsp13a_1", "qsp13a_2", "qsp13a_3"]] = [0.0, 100.0, 0.0]
    out.loc[loss_choice.eq(3), ["qsp13a_1", "qsp13a_2", "qsp13a_3"]] = [0.0, 0.0, 100.0]
    return pd.DataFrame(
        {
            "windfall_gain_save_invest_share": out["qsp12a_1"],
            "windfall_gain_spend_donate_share": out["qsp12a_2"],
            "windfall_gain_pay_debt_share": out["qsp12a_3"],
            "windfall_loss_reduce_spending_share": out["qsp13a_1"],
            "windfall_loss_reduce_saving_share": out["qsp13a_2"],
            "windfall_loss_increase_borrowing_share": out["qsp13a_3"],
        },
        index=frame.index,
    )


def expected_spending_growth_from_density(row: pd.Series) -> float:
    probabilities = pd.to_numeric(row[list(EXPECTED_SPENDING_COLUMNS)], errors="coerce").to_numpy(dtype=float)
    return generalized_beta_density_mean(probabilities)


def generalized_beta_density_mean(probabilities_descending: np.ndarray) -> float:
    probabilities = np.nan_to_num(np.asarray(probabilities_descending, dtype=float), nan=0.0)
    total = float(probabilities.sum())
    if total <= 0:
        return np.nan
    probabilities = probabilities / total
    # Ascending bins: <=-12, -12..-8, -8..-4, -4..-2, -2..0, 0..2, 2..4, 4..8, 8..12, >=12.
    probabilities_ascending = probabilities[::-1]
    cache_key = tuple(float(value) for value in np.round(probabilities_ascending, 4))
    cached = _GENERALIZED_BETA_CACHE.get(cache_key)
    if cached is not None:
        return cached
    bounds = np.array([-38.0, -12.0, -8.0, -4.0, -2.0, 0.0, 2.0, 4.0, 8.0, 12.0, 38.0], dtype=float)
    if int(np.count_nonzero(probabilities_ascending > 0)) == 1:
        index = int(np.argmax(probabilities_ascending))
        out = float((bounds[index] + bounds[index + 1]) / 2.0)
        _GENERALIZED_BETA_CACHE[cache_key] = out
        return out
    cumulative = np.cumsum(probabilities_ascending)[:-1]
    try:
        from scipy.optimize import minimize
        from scipy.special import betainc

        z = (bounds[1:-1] - bounds[0]) / (bounds[-1] - bounds[0])

        def objective(theta: np.ndarray) -> float:
            alpha, beta = np.exp(theta)
            predicted = betainc(alpha, beta, z)
            return float(np.mean(np.square(predicted - cumulative)))

        midpoint = (bounds[:-1] + bounds[1:]) / 2.0
        raw_mean = float(np.dot(probabilities_ascending, midpoint))
        raw_mean_01 = float(np.clip((raw_mean - bounds[0]) / (bounds[-1] - bounds[0]), 0.01, 0.99))
        start = np.log(np.array([2.0 * raw_mean_01 + 0.2, 2.0 * (1.0 - raw_mean_01) + 0.2]))
        result = minimize(objective, start, method="Nelder-Mead", options={"maxiter": 80, "xatol": 1e-4, "fatol": 1e-7})
        alpha, beta = np.exp(result.x)
        out = float(bounds[0] + (bounds[-1] - bounds[0]) * alpha / (alpha + beta))
        _GENERALIZED_BETA_CACHE[cache_key] = out
        return out
    except Exception:
        midpoint = (bounds[:-1] + bounds[1:]) / 2.0
        out = float(np.dot(probabilities_ascending, midpoint))
        _GENERALIZED_BETA_CACHE[cache_key] = out
        return out


def normalize_core_panel(raw: pd.DataFrame) -> pd.DataFrame:
    required = {"sce_raw_userid", "sce_raw_date", "survey_date", "weight", *REGRESSORS, "income_group", "liquid_wealth_group", "age_group"}
    missing = sorted(required - set(raw.columns))
    if missing:
        raise ValueError(f"Core panel missing required columns: {', '.join(missing)}")
    frame = raw.copy()
    frame["userid"] = frame["sce_raw_userid"].astype(str).str.replace(r"\.0$", "", regex=True)
    frame["core_wave"] = pd.to_numeric(frame["sce_raw_date"], errors="coerce").astype("Int64")
    frame["core_wave_date"] = pd.to_datetime(frame["survey_date"], errors="coerce").dt.to_period("M").dt.to_timestamp()
    frame["weight"] = pd.to_numeric(frame["weight"], errors="coerce")
    for column in REGRESSORS:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame.dropna(subset=["userid", "core_wave", "core_wave_date", "weight", *REGRESSORS]).reset_index(drop=True)


def join_spending_to_core_panel(spending: pd.DataFrame, core: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    pieces: list[pd.DataFrame] = []
    core_by_user = {str(userid): group.sort_values("core_wave_date") for userid, group in core.groupby("userid", dropna=False)}
    for userid, spending_group in spending.groupby("userid", dropna=False):
        core_group = core_by_user.get(str(userid))
        if core_group is None or core_group.empty:
            continue
        pieces.append(
            pd.merge_asof(
                spending_group.sort_values("spending_wave_date"),
                core_group,
                left_on="spending_wave_date",
                right_on="core_wave_date",
                direction="backward",
                tolerance=pd.Timedelta(days=92),
            )
        )
    joined = pd.concat(pieces, ignore_index=True) if pieces else pd.DataFrame()
    if "userid_x" in joined:
        joined["userid"] = joined["userid_x"]
    elif "userid_y" in joined:
        joined["userid"] = joined["userid_y"]
    joined = joined.dropna(subset=["core_wave", "weight", *REGRESSORS]).copy()
    joined["belief_lag_days"] = (joined["spending_wave_date"] - joined["core_wave_date"]).dt.days.astype(int)
    joined["belief_lag_months"] = joined["belief_lag_days"] / 30.4375
    columns = [
        "userid",
        "spending_wave",
        "spending_wave_date",
        "core_wave",
        "core_wave_date",
        "belief_lag_days",
        "belief_lag_months",
        "weight",
        "income_group",
        "liquid_wealth_group",
        "age_group",
        "reported_total_spending_growth_pct",
        "expected_total_spending_growth_pct",
        *REGRESSORS,
        "windfall_gain_save_invest_share",
        "windfall_gain_spend_donate_share",
        "windfall_gain_pay_debt_share",
        "windfall_loss_reduce_spending_share",
        "windfall_loss_reduce_saving_share",
        "windfall_loss_increase_borrowing_share",
    ]
    out = joined[columns].sort_values(["spending_wave", "userid"]).reset_index(drop=True)
    report = {
        "spending_rows": int(spending.shape[0]),
        "matched_rows": int(out.shape[0]),
        "matched_share": float(out.shape[0] / max(spending.shape[0], 1)),
        "lag_days": {
            "min": float(out["belief_lag_days"].min()) if not out.empty else None,
            "median": float(out["belief_lag_days"].median()) if not out.empty else None,
            "max": float(out["belief_lag_days"].max()) if not out.empty else None,
        },
        "matches_by_wave": {str(int(index)): int(value) for index, value in out.groupby("spending_wave")["userid"].count().items()},
    }
    return out, report


def read_chart_data(path: Path) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name="Data", header=None)
    metadata = raw.iloc[:6, :].reset_index(drop=True)
    rows: list[dict[str, Any]] = []
    for column_index in range(1, raw.shape[1]):
        first = str(metadata.iat[0, column_index] or "")
        second = str(metadata.iat[1, column_index] or "")
        third = str(metadata.iat[2, column_index] or "")
        title = str(metadata.iat[3, column_index] or "")
        subgroup = str(metadata.iat[4, column_index] or "")
        detail = str(metadata.iat[5, column_index] or "")
        values = raw.iloc[6:, [0, column_index]].dropna(subset=[0])
        for _, value_row in values.iterrows():
            rows.append(
                {
                    "date": int(value_row.iloc[0]),
                    "value": pd.to_numeric(value_row.iloc[1], errors="coerce"),
                    "first_level": first,
                    "second_level": second,
                    "third_level": third,
                    "chart_title": title,
                    "subgroup": subgroup,
                    "detail": detail,
                }
            )
    return pd.DataFrame(rows).dropna(subset=["value"]).reset_index(drop=True)


def verify_expected_spending_chart(spending: pd.DataFrame, chart: pd.DataFrame, core: pd.DataFrame | None = None, *, max_abs_error_pp: float) -> ChartReproduction:
    expected = chart[
        chart["first_level"].str.lower().eq("expectations")
        & chart["second_level"].str.lower().eq("change in spending")
        & chart["third_level"].str.lower().eq("total spending")
        & chart["chart_title"].str.lower().eq("total spending")
        & chart["subgroup"].str.lower().eq("overview")
    ][["date", "value"]].rename(columns={"value": "published"})
    frame = spending.copy()
    if core is not None:
        same_wave_weights = core[["userid", "core_wave", "weight"]].rename(columns={"core_wave": "spending_wave"})
        frame = frame.merge(same_wave_weights, on=["userid", "spending_wave"], how="left")
    rows = []
    for wave, group in frame.groupby("spending_wave"):
        if int(wave) < 201508:
            continue
        predicted = weighted_quantile(group["expected_total_spending_growth_pct"], group["weight"] if "weight" in group else None, 0.5)
        rows.append({"date": int(wave), "microdata": predicted})
    return _chart_check("expected_total_spending_growth_pct", pd.DataFrame(rows), expected, max_abs_error_pp=max_abs_error_pp)


def verify_reported_spending_chart(spending: pd.DataFrame, core: pd.DataFrame, chart: pd.DataFrame, *, max_abs_error_pp: float) -> ChartReproduction:
    reported = chart[
        chart["first_level"].str.lower().eq("experiences")
        & chart["second_level"].str.lower().eq("change in spending")
        & chart["chart_title"].str.lower().eq("change in spending")
        & chart["subgroup"].str.lower().eq("overview")
    ][["date", "value"]].rename(columns={"value": "published"})
    same_wave_weights = core[["userid", "core_wave", "weight"]].rename(columns={"core_wave": "spending_wave"})
    weighted = spending.merge(same_wave_weights, on=["userid", "spending_wave"], how="left")
    rows = []
    for wave, group in weighted.groupby("spending_wave"):
        predicted = grouped_interpolated_median(group["reported_total_spending_growth_pct"], group["weight"])
        rows.append({"date": int(wave), "microdata": predicted})
    return _chart_check("reported_total_spending_growth_pct", pd.DataFrame(rows), reported, max_abs_error_pp=max_abs_error_pp)


def _chart_check(series: str, microdata: pd.DataFrame, published: pd.DataFrame, *, max_abs_error_pp: float) -> ChartReproduction:
    joined = microdata.merge(published, on="date", how="inner")
    joined["abs_error_pp"] = (joined["microdata"] - joined["published"]).abs()
    max_error = float(joined["abs_error_pp"].max()) if not joined.empty else float("inf")
    rows = [
        {
            "date": int(row["date"]),
            "microdata": float(row["microdata"]),
            "published": float(row["published"]),
            "abs_error_pp": float(row["abs_error_pp"]),
        }
        for _, row in joined.sort_values("date").iterrows()
    ]
    return ChartReproduction(series=series, max_abs_error_pp=max_error, passed=bool(max_error <= max_abs_error_pp), rows=rows)


def grouped_interpolated_median(values: pd.Series, weights: pd.Series | None = None) -> float:
    series = pd.to_numeric(values, errors="coerce")
    mask = series.notna()
    series = series[mask]
    if weights is None:
        weight_series = pd.Series(1.0, index=series.index)
    else:
        weight_series = pd.to_numeric(weights, errors="coerce")[mask].fillna(0.0)
    positive = weight_series.gt(0)
    series = series[positive]
    weight_series = weight_series[positive]
    if series.empty:
        return float("nan")
    grouped = weight_series.groupby(series).sum().sort_index()
    total = float(grouped.sum())
    half = 0.5 * total
    cumulative = 0.0
    for value, weight in grouped.items():
        if cumulative + float(weight) >= half:
            return float((float(value) - 0.5) + (half - cumulative) / float(weight))
        cumulative += float(weight)
    return float(grouped.index[-1])


def weighted_quantile(values: pd.Series, weights: pd.Series | None, q: float) -> float:
    series = pd.to_numeric(values, errors="coerce")
    mask = series.notna()
    series = series[mask]
    if weights is None:
        weight_series = pd.Series(1.0, index=series.index)
    else:
        weight_series = pd.to_numeric(weights, errors="coerce")[mask].fillna(0.0)
    positive = weight_series.gt(0)
    series = series[positive]
    weight_series = weight_series[positive]
    if series.empty:
        return float("nan")
    order = np.argsort(series.to_numpy(dtype=float))
    values_array = series.to_numpy(dtype=float)[order]
    weights_array = weight_series.to_numpy(dtype=float)[order]
    cumulative = np.cumsum(weights_array) - 0.5 * weights_array
    cumulative /= weights_array.sum()
    return float(np.interp(q, cumulative, values_array))


def _wave_to_date(series: pd.Series) -> pd.Series:
    text = series.astype(str).str.replace(r"\.0$", "", regex=True)
    parsed = pd.to_datetime(text + "01", format="%Y%m%d", errors="coerce")
    return parsed.dt.to_period("M").dt.to_timestamp()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _safe_relative(path: Path) -> str:
    resolved = path.resolve()
    root = PROJECT_ROOT.resolve()
    return str(resolved.relative_to(root)) if resolved.is_relative_to(root) else str(resolved)


def write_blocker(message: str) -> None:
    path = PROJECT_ROOT / "work" / "codex_briefs" / "empirical_bridge_v3_blockers.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    previous = path.read_text(encoding="utf-8") if path.exists() else "# Empirical Bridge v3 Blockers\n\n"
    entry = f"## {datetime.now(timezone.utc).isoformat()}\n\n{message}\n\n"
    path.write_text(previous + entry, encoding="utf-8")


def _jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_jsonable(item) for item in value]
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (pd.Timestamp,)):
        return value.isoformat()
    if pd.isna(value):
        return None
    return value


if __name__ == "__main__":
    raise SystemExit(main())
