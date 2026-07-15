from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
TARGET_SPECS = {
    "expected_inflation_1y": {"lower": -5.0, "upper": 20.0},
    "expected_unemployment_higher_prob": {"lower": 0.0, "upper": 100.0},
    "expected_real_income_growth": {"lower": -20.0, "upper": 20.0},
}
DEFAULT_INPUT = PROJECT_ROOT / "work" / "persona_beliefs" / "sce_raw" / "frbny-sce-public-microdata-latest.xlsx"
DEFAULT_RAW_DIR = PROJECT_ROOT / "work" / "persona_beliefs" / "sce_raw"
DEFAULT_OUTPUT = PROJECT_ROOT / "work" / "persona_beliefs" / "sce_real_microdata.csv"
DEFAULT_MANIFEST = PROJECT_ROOT / "work" / "persona_beliefs" / "sce_real_microdata_manifest.json"
SCE_REAL_TARGET_FIELDS = (
    "expected_inflation_1y",
    "expected_unemployment_higher_prob",
    "expected_real_income_growth",
)
SCE_REQUIRED_COLUMNS = (
    "date",
    "userid",
    "weight",
    "Q4new",
    "Q9_cent50",
    "Q25v2part2",
)
SCE_OPTIONAL_COLUMNS = (
    "Q13new",
    "Q32",
    "_AGE_CAT",
    "Q33",
    "Q36",
    "D6",
    "_REGION_CAT",
    "Q10_1",
    "Q10_2",
    "Q43",
    "_EDU_CAT",
    "_HH_INC_CAT",
)
SCE_OUTPUT_COLUMNS = (
    "respondent_id",
    "survey_source",
    "survey_date",
    "weight",
    "age_group",
    "income_group",
    "education_group",
    "gender",
    "region",
    "employment_status",
    "homeownership",
    "liquid_wealth_group",
    "actual_expected_inflation_1y",
    "actual_expected_unemployment_higher_prob",
    "actual_expected_real_income_growth",
    "sce_nominal_income_growth",
    "sce_question_unemployment_higher_prob",
    "sce_personal_job_loss_probability_1y",
    "sce_raw_userid",
    "sce_raw_date",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert public NY Fed SCE microdata into the repo persona schema.")
    parser.add_argument("--input-xlsx", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--raw-dir", type=Path, default=DEFAULT_RAW_DIR)
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--start-date", default=None, help="Optional YYYYMM or YYYY-MM lower survey-date bound.")
    parser.add_argument("--end-date", default=None, help="Optional YYYYMM or YYYY-MM upper survey-date bound.")
    parser.add_argument("--max-rows", type=int, default=0, help="Optional debug cap after header row; 0 means all rows.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = convert_sce_microdata(
        input_xlsx=args.input_xlsx,
        raw_dir=args.raw_dir,
        output_csv=args.output_csv,
        manifest_path=args.manifest,
        start_date=args.start_date,
        end_date=args.end_date,
        max_rows=args.max_rows,
    )
    print(result["output_csv"])
    print(result["manifest"])
    return 0


def convert_sce_microdata(
    *,
    input_xlsx: Path,
    raw_dir: Path | None,
    output_csv: Path,
    manifest_path: Path,
    start_date: str | None = None,
    end_date: str | None = None,
    max_rows: int = 0,
) -> dict[str, Any]:
    if not input_xlsx.exists():
        raise FileNotFoundError(f"SCE workbook not found: {input_xlsx}")
    raw = read_sce_workbook(input_xlsx, max_rows=max_rows)
    demographic_lookup = read_sce_demographic_lookup(raw_dir=raw_dir, primary=input_xlsx, max_rows=max_rows)
    converted = normalize_sce_raw_frame(raw, start_date=start_date, end_date=end_date, demographic_lookup=demographic_lookup)
    if converted.empty:
        raise ValueError("No scoreable SCE rows after conversion")
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    converted.to_csv(output_csv, index=False)
    manifest = {
        "schema_version": "sce_real_microdata_conversion_v2",
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "source_file": {
            "path": _safe_relative(input_xlsx),
            "sha256": _file_sha256(input_xlsx),
            "byte_size": int(input_xlsx.stat().st_size),
        },
        "demographic_source_files": [
            _safe_relative(path) for path in _sce_workbooks(raw_dir=raw_dir, primary=input_xlsx)
        ],
        "output_csv": _safe_relative(output_csv),
        "raw_rows": int(raw.shape[0]),
        "converted_rows": int(converted.shape[0]),
        "respondent_count": int(converted["respondent_id"].nunique()),
        "survey_date_min": str(converted["survey_date"].min()),
        "survey_date_max": str(converted["survey_date"].max()),
        "target_fields": list(SCE_REAL_TARGET_FIELDS),
        "target_mapping": {
            "expected_inflation_1y": "Q9_cent50 density median one-year inflation expectation",
            "expected_unemployment_higher_prob": "Q4new percent chance U.S. unemployment will be higher in 12 months",
            "expected_real_income_growth": "Q25v2part2 nominal household income growth minus respondent Q9_cent50 inflation expectation",
        },
        "household_state_mapping": {
            "sce_personal_job_loss_probability_1y": (
                "Q13new percent chance the respondent loses the current/main job "
                "during the next 12 months; missing when not asked"
            ),
        },
        "demographic_handling": (
            "Demographic fields are forward/back-filled within userid before wave filtering, then mapped from "
            "public SCE category columns where exact question fields are sparse."
        ),
        "attribution": (
            "Source: Survey of Consumer Expectations, Federal Reserve Bank of New York. "
            "The repo uses public-use SCE responses as held-out targets; prompt cards must not include actual_* columns."
        ),
    }
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True), encoding="utf-8")
    return {
        "output_csv": str(output_csv),
        "manifest": str(manifest_path),
        "converted_rows": int(converted.shape[0]),
        "respondent_count": int(converted["respondent_id"].nunique()),
    }


def read_sce_workbook(path: Path, *, max_rows: int = 0, columns: tuple[str, ...] | list[str] | None = None) -> pd.DataFrame:
    wanted = list(dict.fromkeys(columns if columns is not None else [*SCE_REQUIRED_COLUMNS, *SCE_OPTIONAL_COLUMNS]))
    try:
        import python_calamine  # noqa: F401

        frame = pd.read_excel(
            path,
            sheet_name="Data",
            header=1,
            usecols=lambda column: column in set(wanted),
            nrows=max_rows if max_rows > 0 else None,
            engine="calamine",
        )
        missing = sorted(set(SCE_REQUIRED_COLUMNS if columns is None else ("userid",)) - set(frame.columns))
        if missing:
            raise ValueError(f"SCE workbook missing required columns: {', '.join(missing)}")
        return frame
    except ImportError:
        pass
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    header_row_index, header = _find_header_row(worksheet)
    indices = {name: header.index(name) for name in wanted if name in header}
    missing = sorted(set(SCE_REQUIRED_COLUMNS if columns is None else ("userid",)) - set(indices))
    if missing:
        raise ValueError(f"SCE workbook missing required columns: {', '.join(missing)}")
    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=1):
        if max_rows > 0 and row_number > max_rows:
            break
        rows.append({name: row[index] if index < len(row) else None for name, index in indices.items()})
    return pd.DataFrame(rows)


def read_sce_demographic_lookup(*, raw_dir: Path | None, primary: Path, max_rows: int = 0) -> pd.DataFrame:
    workbooks = _sce_workbooks(raw_dir=raw_dir, primary=primary)
    frames = [
        read_sce_workbook(path, max_rows=max_rows, columns=["userid", "date", *SCE_OPTIONAL_COLUMNS])
        for path in workbooks
    ]
    if not frames:
        return pd.DataFrame()
    raw = pd.concat(frames, ignore_index=True)
    if raw.empty or "userid" not in raw:
        return pd.DataFrame()
    raw["_survey_timestamp"] = pd.to_datetime(_normalize_sce_month(raw.get("date", pd.Series(index=raw.index))), errors="coerce")
    raw = raw.sort_values(["userid", "_survey_timestamp"])
    rows: list[dict[str, Any]] = []
    for userid, group in raw.groupby("userid", dropna=False):
        if pd.isna(userid):
            continue
        row: dict[str, Any] = {"userid": userid}
        for column in SCE_OPTIONAL_COLUMNS:
            if column not in group:
                continue
            values = group[column].replace("", pd.NA).dropna()
            if not values.empty:
                row[column] = values.iloc[-1]
        rows.append(row)
    return pd.DataFrame(rows).set_index("userid") if rows else pd.DataFrame()


def _sce_workbooks(*, raw_dir: Path | None, primary: Path) -> list[Path]:
    if raw_dir is None or not raw_dir.exists():
        return [primary]
    paths = sorted(path for path in raw_dir.glob("*.xlsx") if path.is_file())
    if primary not in paths:
        paths.append(primary)
    return sorted(set(paths), key=lambda path: path.name)


def normalize_sce_raw_frame(
    raw: pd.DataFrame,
    *,
    start_date: str | None = None,
    end_date: str | None = None,
    demographic_lookup: pd.DataFrame | None = None,
) -> pd.DataFrame:
    frame = raw.copy()
    frame["survey_date"] = _normalize_sce_month(frame["date"])
    frame = frame.dropna(subset=["userid", "survey_date"]).copy()
    frame["survey_timestamp"] = pd.to_datetime(frame["survey_date"], errors="coerce")
    frame = frame.dropna(subset=["survey_timestamp"]).sort_values(["userid", "survey_timestamp"])
    if demographic_lookup is not None and not demographic_lookup.empty:
        for column in SCE_OPTIONAL_COLUMNS:
            if column in demographic_lookup:
                if column not in frame:
                    frame[column] = pd.NA
                frame[column] = frame[column].where(
                    frame[column].notna() & frame[column].astype(str).ne(""),
                    frame["userid"].map(demographic_lookup[column]),
                )
    for column in ["Q32", "_AGE_CAT", "Q33", "Q36", "D6", "_REGION_CAT", "Q43", "_EDU_CAT", "_HH_INC_CAT"]:
        if column in frame:
            frame[column] = frame.groupby("userid", dropna=False)[column].transform(lambda series: series.ffill().bfill())
    if start_date:
        frame = frame[frame["survey_timestamp"].ge(_parse_month_bound(start_date, end=False))]
    if end_date:
        frame = frame[frame["survey_timestamp"].le(_parse_month_bound(end_date, end=True))]

    inflation = pd.to_numeric(frame["Q9_cent50"], errors="coerce")
    unemployment_higher = pd.to_numeric(frame["Q4new"], errors="coerce")
    personal_job_loss = pd.to_numeric(_optional_column(frame, "Q13new"), errors="coerce")
    nominal_income = pd.to_numeric(frame["Q25v2part2"], errors="coerce")
    weight = pd.to_numeric(frame["weight"], errors="coerce").fillna(0.0).clip(lower=0.0)
    out = pd.DataFrame(
        {
            "respondent_id": "sce_user_" + frame["userid"].astype(str).str.replace(r"\.0$", "", regex=True),
            "survey_source": "ny_fed_sce_real_microdata",
            "survey_date": frame["survey_date"].astype(str),
            "weight": weight,
            "age_group": _age_group(frame),
            "income_group": _income_group(frame, "_HH_INC_CAT"),
            "education_group": _education_group(frame, "_EDU_CAT"),
            "gender": _gender(frame, "Q33"),
            "region": _clean_category(_optional_column(frame, "_REGION_CAT"), default="unknown"),
            "employment_status": _employment_status(frame),
            "homeownership": _homeownership(frame, "Q43"),
            "liquid_wealth_group": _income_group(frame, "_HH_INC_CAT"),
            "actual_expected_inflation_1y": inflation.clip(
                TARGET_SPECS["expected_inflation_1y"]["lower"],
                TARGET_SPECS["expected_inflation_1y"]["upper"],
            ),
            "actual_expected_unemployment_higher_prob": unemployment_higher.clip(
                TARGET_SPECS["expected_unemployment_higher_prob"]["lower"],
                TARGET_SPECS["expected_unemployment_higher_prob"]["upper"],
            ),
            "actual_expected_real_income_growth": (nominal_income - inflation).clip(
                TARGET_SPECS["expected_real_income_growth"]["lower"],
                TARGET_SPECS["expected_real_income_growth"]["upper"],
            ),
            "sce_nominal_income_growth": nominal_income,
            "sce_question_unemployment_higher_prob": unemployment_higher,
            "sce_personal_job_loss_probability_1y": personal_job_loss.clip(0.0, 100.0),
            "sce_raw_userid": frame["userid"].astype(str).str.replace(r"\.0$", "", regex=True),
            "sce_raw_date": frame["date"],
        }
    )
    out = out.dropna(
        subset=[
            "actual_expected_inflation_1y",
            "actual_expected_unemployment_higher_prob",
            "actual_expected_real_income_growth",
        ]
    ).copy()
    out = out[out["weight"].gt(0)].copy()
    out["weight"] = out.groupby("survey_date")["weight"].transform(lambda series: series / float(series.sum()) if float(series.sum()) > 0 else 1.0 / len(series))
    return out[list(SCE_OUTPUT_COLUMNS)].reset_index(drop=True)


def _find_header_row(worksheet: Any) -> tuple[int, list[str]]:
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        values = [str(value) if value is not None else "" for value in row]
        if {"date", "userid", "weight"}.issubset(set(values)):
            return row_index, values
    raise ValueError("Could not find SCE header row containing date, userid, and weight")


def _normalize_sce_month(series: pd.Series) -> pd.Series:
    text = series.astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    yyyymm = text.str.fullmatch(r"\d{6}")
    out = pd.Series(pd.NA, index=series.index, dtype="object")
    out.loc[yyyymm] = text.loc[yyyymm].str.slice(0, 4) + "-" + text.loc[yyyymm].str.slice(4, 6) + "-01"
    if (~yyyymm).any():
        parsed = pd.to_datetime(text.loc[~yyyymm], errors="coerce")
        out.loc[~yyyymm] = parsed.dt.to_period("M").dt.to_timestamp().dt.date.astype(str).where(parsed.notna(), pd.NA)
    return out


def _parse_month_bound(value: str, *, end: bool) -> pd.Timestamp:
    text = str(value).strip()
    if len(text) == 6 and text.isdigit():
        ts = pd.Timestamp(f"{text[:4]}-{text[4:6]}-01")
    else:
        ts = pd.Timestamp(text if len(text) > 7 else f"{text}-01")
    if end:
        return ts + pd.offsets.MonthEnd(0)
    return ts


def _age_group(frame: pd.DataFrame) -> pd.Series:
    numeric = pd.to_numeric(_optional_column(frame, "Q32"), errors="coerce")
    exact = pd.Series(
        np.select([numeric < 35, numeric < 55, numeric >= 55], ["18_34", "35_54", "55_plus"], default=pd.NA),
        index=frame.index,
        dtype="object",
    )
    category = _clean_category(_optional_column(frame, "_AGE_CAT"), default="unknown")
    category = category.replace({"under_40": "18_34", "40_to_60": "35_54", "over_60": "55_plus"})
    return exact.where(exact.notna(), category).fillna("unknown")


def _income_group(frame: pd.DataFrame, column: str) -> pd.Series:
    category = _clean_category(_optional_column(frame, column), default="unknown")
    return category.replace({"under_50k": "low", "50k_to_100k": "middle", "over_100k": "high", "": "unknown"})


def _education_group(frame: pd.DataFrame, column: str) -> pd.Series:
    category = _clean_category(_optional_column(frame, column), default="unknown")
    return category.replace({"high_school": "high_school_or_less", "some_college": "some_college", "college": "college_plus"})


def _gender(frame: pd.DataFrame, column: str) -> pd.Series:
    series = _optional_column(frame, column)
    numeric = pd.to_numeric(series, errors="coerce")
    return pd.Series(np.select([numeric.eq(1), numeric.eq(2)], ["female", "male"], default="unknown"), index=series.index)


def _employment_status(frame: pd.DataFrame) -> pd.Series:
    q10_1 = pd.to_numeric(_optional_column(frame, "Q10_1"), errors="coerce")
    q10_2 = pd.to_numeric(_optional_column(frame, "Q10_2"), errors="coerce")
    return pd.Series(
        np.select([q10_2.eq(1), q10_1.eq(1)], ["unemployed", "employed"], default="unknown"),
        index=frame.index,
    )


def _homeownership(frame: pd.DataFrame, column: str) -> pd.Series:
    series = _optional_column(frame, column)
    numeric = pd.to_numeric(series, errors="coerce")
    return pd.Series(np.select([numeric.eq(1), numeric.eq(2)], ["owner", "renter"], default="unknown"), index=series.index)


def _optional_column(frame: pd.DataFrame, column: str) -> pd.Series:
    if column in frame:
        return frame[column]
    return pd.Series(pd.NA, index=frame.index, dtype="object")


def _clean_category(series: pd.Series, *, default: str) -> pd.Series:
    cleaned = series.fillna(default).astype(str).str.strip().str.lower().str.replace(" ", "_").str.replace("-", "_")
    return cleaned.replace({"": default, "nan": default, "none": default})


def _safe_relative(path: Path) -> str:
    resolved = path.resolve()
    cwd = Path.cwd().resolve()
    return str(resolved.relative_to(cwd)) if resolved.is_relative_to(cwd) else str(resolved)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


if __name__ == "__main__":
    raise SystemExit(main())
